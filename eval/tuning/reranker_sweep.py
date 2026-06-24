"""
reranker_sweep.py — tune RAG retrieval dials against the retrieval_labels ground truth.

Ground truth = the labelled relevant PAGE RANGES in eval/datasets/retrieval_labels.jsonl
(derived from the indexed chunk text, independent of the retriever). A retrieved chunk
counts as relevant if its page range overlaps a labelled range. For each config we
report mean precision@k / recall@k / MRR / nDCG + mean latency/query, so you can see
whether the cross-encoder reranker earns its cost and which RERANK_CANDIDATES is best.

No LLM — local embedder + reranker only (CPU-slow but free). Measures in-process;
NEVER writes config. Needs the Chroma index (rag/chroma_store).

    python eval/tuning/reranker_sweep.py
"""

import json
import math
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "rag"))

import retriever as R
import openpyxl
from openpyxl.styles import Font, PatternFill

LABELS = ROOT / "eval" / "datasets" / "retrieval_labels.jsonl"
OUT = ROOT / "eval" / "results" / "tuning"

# (label, rerank?, candidate count). candidate count is ignored when rerank is off.
CONFIGS = [
    ("cosine-only (no rerank)", False, None),
    ("rerank, cand=8", True, 8),
    ("rerank, cand=12", True, 12),
    ("rerank, cand=20", True, 20),
]


def _retrieve(query, mvc, k, rerank, candidates):
    """Mirror retriever's two functions, but with the dials parameterised."""
    n = (candidates or k) if rerank else k
    where = {"mvc_code": mvc} if mvc else {"doc_type": "safety"}
    qv = R.get_embedding_model().embed_query(query)
    res = R._collection().query(query_embeddings=[qv], n_results=n, where=where)
    cands = R._format(res)
    out = R._rerank(query, cands, k) if rerank else cands[:k]
    return [{"source_file": c["metadata"].get("source_file"),
             "page_start": c["metadata"].get("page_start"),
             "page_end": c["metadata"].get("page_end")} for c in out]


def _overlaps(c, r):
    return (c["source_file"] == r["source_file"]
            and not (c["page_end"] < r["page_start"] or c["page_start"] > r["page_end"]))


def _metrics(retrieved, relevant):
    k = len(retrieved)
    hits = [1 if any(_overlaps(c, r) for r in relevant) else 0 for c in retrieved]
    precision = sum(hits) / k if k else 0.0
    recall = (sum(1 for r in relevant if any(_overlaps(c, r) for c in retrieved)) / len(relevant)
              if relevant else 0.0)
    mrr = next((1 / (i + 1) for i, h in enumerate(hits) if h), 0.0)
    dcg = sum(h / math.log2(i + 2) for i, h in enumerate(hits))
    ideal = sum(1 / math.log2(i + 2) for i in range(min(len(relevant), k))) if k else 0.0
    ndcg = dcg / ideal if ideal else 0.0
    return precision, recall, mrr, ndcg


def run():
    rows = [json.loads(l) for l in LABELS.read_text().splitlines() if l.strip()]
    print(f"reranker_sweep over {len(rows)} retrieval_labels examples\n")
    results = []
    for label, rerank, cand in CONFIGS:
        agg = [0.0, 0.0, 0.0, 0.0]; t0 = time.perf_counter()
        for ex in rows:
            inp, ref = ex["inputs"], ex["reference"]
            retrieved = _retrieve(inp["query"], inp.get("mvc_code"), inp.get("k", 5), rerank, cand)
            p, r, m, n = _metrics(retrieved, ref["relevant"])
            for i, v in enumerate((p, r, m, n)):
                agg[i] += v
        nrows = len(rows)
        lat = (time.perf_counter() - t0) / nrows
        rec = {"config": label, "precision@k": agg[0] / nrows, "recall@k": agg[1] / nrows,
               "mrr": agg[2] / nrows, "ndcg": agg[3] / nrows, "latency_s": lat}
        results.append(rec)
        print(f"  {label:26}  P={rec['precision@k']:.2f} R={rec['recall@k']:.2f} "
              f"MRR={rec['mrr']:.2f} nDCG={rec['ndcg']:.2f}  {lat:.2f}s/query")

    best = max(results, key=lambda x: (round(x["recall@k"], 3), x["ndcg"]))
    print(f"\nBest recall@k: '{best['config']}' (recall={best['recall@k']:.2f}, nDCG={best['ndcg']:.2f})")
    print("Current dial: RERANK_CANDIDATES =", R.RERANK_CANDIDATES,
          "| report only — apply via config after review (record in TUNING_LOG.md).")

    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / f"reranker_sweep_{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "reranker_sweep"
    cols = ["config", "precision@k", "recall@k", "mrr", "ndcg", "latency_s"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for rec in results:
        ws.append([rec["config"]] + [round(rec[c], 3) for c in cols[1:]])
        if rec is best:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor="C6EFCE")
    wb.save(path)
    print(f"\nExcel: {path}")


if __name__ == "__main__":
    run()
