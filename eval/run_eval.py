"""
run_eval.py — Phase 5c runner. Binds each target to its uploaded LangSmith dataset,
runs the evaluators, and emits: a LangSmith Experiment per dataset, a colour-coded
Excel workbook (eval/results/eval_<ts>.xlsx), and a markdown summary.

    python eval/run_eval.py                         # all 6 datasets (default)
    python eval/run_eval.py --dataset routing       # one dataset (substring match)

Prereqs: LANGSMITH_API_KEY + OPENROUTER_API_KEY + GROQ/GOOGLE keys in .env. The
troubleshoot + manage datasets also need the HTTP MCP server up:
    python mcp_server/server.py http
Pass a timestamp via --stamp to keep filenames deterministic (Date.now is fine here).
"""

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "eval"))
sys.path.insert(0, str(ROOT / "eval" / "datasets"))
sys.path.insert(0, str(ROOT))

from datasets.schemas import DATASETS
from evaluators import EVALUATORS
from targets import TARGETS
from thresholds import threshold

from langsmith import aevaluate
import openpyxl
from openpyxl.styles import Font, PatternFill

RESULTS = ROOT / "eval" / "results"
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
HEAD = Font(bold=True)
COLS = ["case_id", "input", "expected", "agent_output", "scores", "correct", "result", "comments"]


def _short(obj, n):
    try:
        return json.dumps(obj, default=str)[:n]
    except Exception:
        return str(obj)[:n]


def _row_verdict(feedbacks):
    """(scores_str, comments_str, pass_bool|None) from a list of EvaluationResults."""
    scored, comments, fails = [], [], 0
    any_scored = False
    for f in feedbacks:
        key = getattr(f, "key", None) or (f.get("key") if isinstance(f, dict) else None)
        score = getattr(f, "score", None) if not isinstance(f, dict) else f.get("score")
        comment = getattr(f, "comment", None) if not isinstance(f, dict) else f.get("comment")
        if score is not None:
            scored.append(f"{key}={score:.2f}" if isinstance(score, float) else f"{key}={score}")
            any_scored = True
            if score < threshold(key):
                fails += 1
        if comment:
            comments.append(f"{key}: {comment}")
    passed = None if not any_scored else (fails == 0)
    return "; ".join(scored), " | ".join(comments)[:500], passed


async def run_dataset(fname, stamp):
    ds_name = DATASETS[fname][1]
    target, evs = TARGETS[fname], EVALUATORS[fname]
    print(f"\n=== {ds_name} ===")
    results = await aevaluate(
        target, data=ds_name, evaluators=evs,
        experiment_prefix=f"{fname.split('.')[0]}-{stamp}", max_concurrency=3)
    rows = []
    async for r in results:
        run, ex = r["run"], r["example"]
        fbs = (r.get("evaluation_results") or {}).get("results", [])
        scores, comments, passed = _row_verdict(fbs)
        out = dict(run.outputs or {}); out.pop("context_text", None)
        rows.append({
            "case_id": (ex.metadata or {}).get("case_id", ex.id),
            "input": _short(ex.inputs, 300),
            "expected": _short(ex.outputs, 400),
            "agent_output": _short(out, 600),
            "scores": scores,
            "correct": "n/a" if passed is None else ("right" if passed else "wrong"),
            "result": "n/a" if passed is None else ("PASS" if passed else "FAIL"),
            "comments": comments,
        })
    name = getattr(results, "experiment_name", ds_name)
    print(f"  {len(rows)} examples — experiment '{name}'")
    return rows


def write_workbook(by_ds, path):
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["dataset", "examples", "pass", "fail", "n/a", "pass_rate"])
    for c in summary[1]:
        c.font = HEAD
    for fname, rows in by_ds.items():
        p = sum(1 for r in rows if r["result"] == "PASS")
        f = sum(1 for r in rows if r["result"] == "FAIL")
        na = sum(1 for r in rows if r["result"] == "n/a")
        graded = p + f
        summary.append([fname, len(rows), p, f, na, f"{(p / graded * 100):.0f}%" if graded else "—"])

        ws = wb.create_sheet(fname.split(".")[0][:31])
        ws.append(COLS)
        for c in ws[1]:
            c.font = HEAD
        for r in rows:
            ws.append([r[c] for c in COLS])
            cell = ws.cell(row=ws.max_row, column=COLS.index("result") + 1)
            cell.fill = GREEN if r["result"] == "PASS" else (RED if r["result"] == "FAIL" else PatternFill())
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(60, max(12, max(len(str(c.value or "")) for c in col) + 2))
    wb.save(path)


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset", default="all", help="substring of a dataset filename, or 'all'")
    ap.add_argument("--stamp", default=None)
    args = ap.parse_args()
    stamp = args.stamp or datetime.now().strftime("%Y%m%d-%H%M%S")

    selected = [f for f in DATASETS if args.dataset == "all" or args.dataset in f]
    if not selected:
        print(f"no dataset matches '{args.dataset}'. options: {list(DATASETS)}"); return

    RESULTS.mkdir(exist_ok=True)
    by_ds = {}
    for fname in selected:
        try:
            by_ds[fname] = await run_dataset(fname, stamp)
        except Exception as e:
            print(f"  !! {fname} failed: {str(e)[:200]}")
    if not by_ds:
        return
    xlsx = RESULTS / f"eval_{stamp}.xlsx"
    write_workbook(by_ds, xlsx)
    print(f"\nExcel: {xlsx}")
    print("Open the Experiments in LangSmith for per-example drill-down.")


if __name__ == "__main__":
    asyncio.run(main())
